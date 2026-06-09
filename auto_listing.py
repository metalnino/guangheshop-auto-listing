# -*- coding: utf-8 -*-
"""
自动上架商品脚本
================
通过直接操作 guangheshop 数据库，实现商品的自动上架。
完整流程：图片上传COS → 创建商品 → 创建SKU → 创建库存批次 → 绑定区域 → 写入日志

数据库：192.168.100.96:13506 / guangheshop
COS桶：mpfamily-1301068541 (ap-shanghai)

使用方式：
    1. 准备 Excel/CSV 商品数据
    2. 配置 COS 密钥（如需上传图片）
    3. python auto_listing.py

作者：auto-generated
创建时间：2026-06-03
"""

import pymysql
from pymysql.cursors import DictCursor
import json
import time
import random
import logging
import os
import csv
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
import shutil
import glob


# ============================================================
# 日志配置
# ============================================================
# 创建 logs 目录并归档历史日志
script_dir = os.path.dirname(__file__)
log_dir = os.path.join(script_dir, 'logs')
os.makedirs(log_dir, exist_ok=True)

# 将主目录下旧的日志文件移入 logs 文件夹中归档
for old_log in glob.glob(os.path.join(script_dir, 'auto_listing_*.log')):
    try:
        shutil.move(old_log, os.path.join(log_dir, os.path.basename(old_log)))
    except Exception:
        pass

log_filename = f'auto_listing_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(log_dir, log_filename), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# 加载外部配置
# ============================================================
from config import DB_CONFIG, COS_CONFIG, DEFAULTS, AI_OUTPUT_DIR


# ============================================================
# 工具函数
# ============================================================
def get_connection():
    """获取数据库连接"""
    return pymysql.connect(**DB_CONFIG)


def generate_goods_sn():
    """
    生成商品编号
    规则：G + 13位时间戳(毫秒) + 3位随机数 → 如 G1778569628445
    """
    timestamp = int(time.time() * 1000)
    rand_suffix = random.randint(100, 999)
    # 取前10位时间戳 + 3位随机数 = 13位数字
    sn = f"G{timestamp}"
    return sn


def calculate_prices(cost_price=None, cost_total=None, wholesale_price=None,
                     wholesale_total=None, quantity=1):
    """
    价格换算规则（文档 §3.2）：
    1. 只给单价 → 总价 = 单价 × qty
    2. 只给总价 → 单价 = 总价 ÷ qty（四舍五入到2位）
    3. 单价、总价都给 → 以总价为准，重算单价
    4. 采购价、批发价各自独立换算
    """
    qty = Decimal(str(quantity))

    # 采购价换算
    if cost_total is not None:
        cost_total = Decimal(str(cost_total))
        cost_price = (cost_total / qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    elif cost_price is not None:
        cost_price = Decimal(str(cost_price))
        cost_total = (cost_price * qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # 批发价换算
    if wholesale_total is not None:
        wholesale_total = Decimal(str(wholesale_total))
        wholesale_price = (wholesale_total / qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    elif wholesale_price is not None:
        wholesale_price = Decimal(str(wholesale_price))
        wholesale_total = (wholesale_price * qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    return cost_price, cost_total, wholesale_price, wholesale_total


# ============================================================
# 路径自动修复
# ============================================================
def resolve_image_path(path):
    """
    解析图片路径，如果文件不存在则尝试在子目录中查找同名文件。
    常见场景：用户填了 .../scene/scene_05.png 但实际是 .../scenes/scene_05.png
    """
    if os.path.exists(path):
        return path

    filename = os.path.basename(path)
    parent = os.path.dirname(os.path.dirname(path))  # 上两级目录

    if os.path.isdir(parent):
        # 在 parent 的所有子目录中搜索同名文件
        for dirpath, dirnames, filenames in os.walk(parent):
            if filename in filenames:
                found = os.path.join(dirpath, filename)
                logger.warning(f"路径自动修复: {path} -> {found}")
                return found

    raise FileNotFoundError(f"图片文件不存在且无法自动修复: {path}")


# ============================================================
# 团长名称 → 用户ID 查询
# ============================================================
_creator_cache = {}  # 名称 -> user_id 缓存，避免重复查询


def resolve_creator_id(creator_name):
    """
    通过团长微信昵称查询对应的 litemall_user.id

    查询逻辑：
        z100h_wx.wx_user.name (LIKE 模糊匹配)
        JOIN guangheshop.litemall_user ON unionid = weixin_unionid

    Args:
        creator_name: 团长微信昵称

    Returns:
        int: 用户ID (creator_id)

    Raises:
        ValueError: 找不到或找到多个匹配
    """
    name = str(creator_name).strip()
    if not name:
        raise ValueError("团长名称不能为空")

    # 缓存命中
    if name in _creator_cache:
        return _creator_cache[name]

    conn = get_connection()
    try:
        cursor = conn.cursor()
        sql = """
            SELECT b.id
            FROM z100h_wx.wx_user a
            JOIN guangheshop.litemall_user b ON a.unionid = b.weixin_unionid
            WHERE a.`name` LIKE %s
        """
        cursor.execute(sql, (f'%{name}%',))
        rows = cursor.fetchall()

        if not rows:
            raise ValueError(f"找不到团长 '{name}'，请确认微信昵称是否正确")
        if len(rows) > 1:
            ids = [r['id'] for r in rows]
            raise ValueError(f"团长名称 '{name}' 匹配到多个用户: {ids}，请使用更精确的名称")

        user_id = rows[0]['id']
        _creator_cache[name] = user_id
        logger.info(f"  ✓ 团长名称解析: '{name}' → user_id={user_id}")
        return user_id
    finally:
        conn.close()


def get_all_creators():
    """
    从数据库查询所有有效的团长（创建者）微信昵称列表，并加入默认的团长昵称
    """
    creators = []
    try:
        conn = get_connection()
        cursor = conn.cursor()
        sql = """
            SELECT DISTINCT a.name 
            FROM z100h_wx.wx_user a 
            JOIN guangheshop.litemall_user b ON a.unionid = b.weixin_unionid
            WHERE a.name IS NOT NULL AND a.name != ''
        """
        cursor.execute(sql)
        creators = [r['name'] for r in cursor.fetchall() if r['name']]
        conn.close()
    except Exception as e:
        logger.warning(f"从数据库查询团长列表失败: {e}，将使用默认团长列表。")

    # 去重并确保吴鑫霞和李琳在列表中
    creators = list(set(creators))
    for default_name in ["吴鑫霞", "李琳"]:
        if default_name not in creators:
            creators.append(default_name)
    return sorted(creators)



# ============================================================
# 城市名称 → 编码映射
# ============================================================
CITY_MAP = {
    # 江苏省
    "南京": (320100, "南京市"), "南京市": (320100, "南京市"),
    "无锡": (320200, "无锡市"), "无锡市": (320200, "无锡市"),
    "徐州": (320300, "徐州市"), "徐州市": (320300, "徐州市"),
    "常州": (320400, "常州市"), "常州市": (320400, "常州市"),
    "苏州": (320500, "苏州市"), "苏州市": (320500, "苏州市"),
    "南通": (320600, "南通市"), "南通市": (320600, "南通市"),
    "连云港": (320700, "连云港市"), "连云港市": (320700, "连云港市"),
    "淮安": (320800, "淮安市"), "淮安市": (320800, "淮安市"),
    "盐城": (320900, "盐城市"), "盐城市": (320900, "盐城市"),
    "扬州": (321000, "扬州市"), "扬州市": (321000, "扬州市"),
    "镇江": (321100, "镇江市"), "镇江市": (321100, "镇江市"),
    "泰州": (321200, "泰州市"), "泰州市": (321200, "泰州市"),
    "宿迁": (321300, "宿迁市"), "宿迁市": (321300, "宿迁市"),
    # 浙江省
    "杭州": (330100, "杭州市"), "杭州市": (330100, "杭州市"),
    "宁波": (330200, "宁波市"), "宁波市": (330200, "宁波市"),
    "温州": (330300, "温州市"), "温州市": (330300, "温州市"),
    "嘉兴": (330400, "嘉兴市"), "嘉兴市": (330400, "嘉兴市"),
    "湖州": (330500, "湖州市"), "湖州市": (330500, "湖州市"),
    "绍兴": (330600, "绍兴市"), "绍兴市": (330600, "绍兴市"),
    "金华": (330700, "金华市"), "金华市": (330700, "金华市"),
    "衢州": (330800, "衢州市"), "衢州市": (330800, "衢州市"),
    "舟山": (330900, "舟山市"), "舟山市": (330900, "舟山市"),
    "台州": (331000, "台州市"), "台州市": (331000, "台州市"),
    "丽水": (331100, "丽水市"), "丽水市": (331100, "丽水市"),
    # 湖北省
    "武汉": (420100, "武汉市"), "武汉市": (420100, "武汉市"),
    "黄石": (420200, "黄石市"), "黄石市": (420200, "黄石市"),
    "十堰": (420300, "十堰市"), "十堰市": (420300, "十堰市"),
    "宜昌": (420500, "宜昌市"), "宜昌市": (420500, "宜昌市"),
    "襄阳": (420600, "襄阳市"), "襄阳市": (420600, "襄阳市"),
    "鄂州": (420700, "鄂州市"), "鄂州市": (420700, "鄂州市"),
    "荆门": (420800, "荆门市"), "荆门市": (420800, "荆门市"),
    "孝感": (420900, "孝感市"), "孝感市": (420900, "孝感市"),
    "荆州": (421000, "荆州市"), "荆州市": (421000, "荆州市"),
    "黄冈": (421100, "黄冈市"), "黄冈市": (421100, "黄冈市"),
    "咸宁": (421200, "咸宁市"), "咸宁市": (421200, "咸宁市"),
    "随州": (421300, "随州市"), "随州市": (421300, "随州市"),
    "恩施": (422800, "恩施州"), "恩施州": (422800, "恩施州"),
    # 广东省
    "广州": (440100, "广州市"), "广州市": (440100, "广州市"),
    "韶关": (440200, "韶关市"), "韶关市": (440200, "韶关市"),
    "深圳": (440300, "深圳市"), "深圳市": (440300, "深圳市"),
    "珠海": (440400, "珠海市"), "珠海市": (440400, "珠海市"),
    "汕头": (440500, "汕头市"), "汕头市": (440500, "汕头市"),
    "佛山": (440600, "佛山市"), "佛山市": (440600, "佛山市"),
    "江门": (440700, "江门市"), "江门市": (440700, "江门市"),
    "湛江": (440800, "湛江市"), "湛江市": (440800, "湛江市"),
    "茂名": (440900, "茂名市"), "茂名市": (440900, "茂名市"),
    "肇庆": (441200, "肇庆市"), "肇庆市": (441200, "肇庆市"),
    "惠州": (441300, "惠州市"), "惠州市": (441300, "惠州市"),
    "梅州": (441400, "梅州市"), "梅州市": (441400, "梅州市"),
    "汕尾": (441500, "汕尾市"), "汕尾市": (441500, "汕尾市"),
    "河源": (441600, "河源市"), "河源市": (441600, "河源市"),
    "阳江": (441700, "阳江市"), "阳江市": (441700, "阳江市"),
    "清远": (441800, "清远市"), "清远市": (441800, "清远市"),
    "东莞": (441900, "东莞市"), "东莞市": (441900, "东莞市"),
    "中山": (442000, "中山市"), "中山市": (442000, "中山市"),
    "潮州": (445100, "潮州市"), "潮州市": (445100, "潮州市"),
    "揭阳": (445200, "揭阳市"), "揭阳市": (445200, "揭阳市"),
    "云浮": (445300, "云浮市"), "云浮市": (445300, "云浮市"),
    # 上海
    "上海": (310100, "上海市"), "上海市": (310100, "上海市"),
}
# 省份快捷方式：写"江苏全部"或"江苏"自动展开为该省所有城市
PROVINCE_SHORTCUTS = {
    "江苏省": ["南京市", "无锡市", "徐州市", "常州市", "苏州市", "南通市",
              "连云港市", "淮安市", "盐城市", "扬州市", "镇江市", "泰州市", "宿迁市"],
    "浙江省": ["杭州市", "宁波市", "温州市", "嘉兴市", "湖州市", "绍兴市",
              "金华市", "衢州市", "舟山市", "台州市", "丽水市"],
    "湖北省": ["武汉市", "黄石市", "十堰市", "宜昌市", "襄阳市", "鄂州市",
              "荆门市", "孝感市", "荆州市", "黄冈市", "咸宁市", "随州市", "恩施州"],
    "广东省": ["广州市", "韶关市", "深圳市", "珠海市", "汕头市", "佛山市",
              "江门市", "湛江市", "茂名市", "肇庆市", "惠州市", "梅州市",
              "汕尾市", "河源市", "阳江市", "清远市", "东莞市", "中山市",
              "潮州市", "揭阳市", "云浮市"],
    "上海市": ["上海市"],
}
# 构建快捷名称映射（支持多种写法）
_SHORTCUT_MAP = {}
for prov, cities in PROVINCE_SHORTCUTS.items():
    short = prov.rstrip("省").rstrip("市")  # 江苏、浙江...
    _SHORTCUT_MAP[prov] = cities              # 江苏省
    _SHORTCUT_MAP[short] = cities             # 江苏
    _SHORTCUT_MAP[short + "全部"] = cities    # 江苏全部
    _SHORTCUT_MAP[prov + "全部"] = cities     # 江苏省全部


def resolve_city_regions(city_names_raw):
    """
    将城市名称字符串（用 | 分隔）解析为 regions 列表

    支持写法：
        - 城市名：  "南京" / "南京市" / "南京|苏州|镇江"
        - 省份快捷：  "江苏全部" / "江苏" / "江苏省"
        - 混合写法：  "江苏全部|武汉|深圳"

    Args:
        city_names_raw: 城市名称字符串，多个用 | 分隔

    Returns:
        list[dict]: [{'code': 320100, 'name': '南京市'}, ...]

    Raises:
        ValueError: 无法识别的城市名称
    """
    if not city_names_raw or not city_names_raw.strip():
        raise ValueError("城市不能为空")

    # 第一步：展开省份快捷方式
    expanded_names = []
    for name in city_names_raw.split('|'):
        name = name.strip()
        if not name:
            continue
        if name in _SHORTCUT_MAP:
            expanded_names.extend(_SHORTCUT_MAP[name])
            logger.info(f"  ✓ 省份快捷展开: '{name}' → {len(_SHORTCUT_MAP[name])}个城市")
        else:
            expanded_names.append(name)

    # 第二步：逐个解析城市名
    regions = []
    seen_codes = set()
    for name in expanded_names:
        name = name.strip()
        if not name:
            continue

        if name in CITY_MAP:
            code, full_name = CITY_MAP[name]
        else:
            # 模糊匹配
            matches = [(k, v) for k, v in CITY_MAP.items() if name in k or k in name]
            if len(matches) >= 1:
                unique = {v[0]: (v[0], v[1]) for _, v in matches}
                if len(unique) == 1:
                    code, full_name = list(unique.values())[0]
                else:
                    options = [v[1] for v in unique.values()]
                    raise ValueError(f"城市 '{name}' 匹配到多个: {options}，请更精确")
            else:
                available = ', '.join(sorted(set(v[1] for v in CITY_MAP.values())))
                raise ValueError(f"无法识别 '{name}'\n可选城市: {available}")

        if code not in seen_codes:
            regions.append({'code': code, 'name': full_name})
            seen_codes.add(code)

    if not regions:
        raise ValueError("至少需要选择 1 个城市")

    logger.info(f"  ✓ 城市解析: {city_names_raw} → {[r['name'] for r in regions]}")
    return regions


# ============================================================
# COS 图片上传
# ============================================================
def upload_image_to_cos(local_path):
    """
    将本地图片上传到腾讯云 COS
    
    Args:
        local_path: 本地图片文件路径

    Returns:
        上传后的 COS URL 字符串
    """
    if not COS_CONFIG.get('upload_to_cos', True):
        logger.info(f"[COS-MOCK] 跳过上传，使用本地路径: {local_path}")
        return local_path

    if not COS_CONFIG['secret_id'] or not COS_CONFIG['secret_key']:
        logger.error("COS 密钥未配置！请在 COS_CONFIG 中填入 secret_id 和 secret_key")
        raise ValueError("COS 密钥未配置")

    try:
        from qcloud_cos import CosConfig, CosS3Client
    except ImportError:
        logger.error("请先安装 cos-python-sdk-v5: pip install cos-python-sdk-v5")
        raise

    config = CosConfig(
        Region=COS_CONFIG['region'],
        SecretId=COS_CONFIG['secret_id'],
        SecretKey=COS_CONFIG['secret_key'],
    )
    client = CosS3Client(config)

    # 生成 COS key：activity/{原始文件名}_{时间戳}.{ext}
    basename = os.path.splitext(os.path.basename(local_path))[0]  # 原始文件名(不含扩展名)
    ext = os.path.splitext(local_path)[1].lower() or '.jpg'
    timestamp = int(time.time() * 1000)
    cos_key = f"{COS_CONFIG['upload_prefix']}{basename}_{timestamp}{ext}"

    logger.info(f"正在上传图片到 COS: {local_path} → {cos_key}")
    client.upload_file(
        Bucket=COS_CONFIG['bucket'],
        LocalFilePath=local_path,
        Key=cos_key,
    )

    url = f"https://{COS_CONFIG['bucket']}.cos.{COS_CONFIG['region']}.myqcloud.com/{cos_key}"
    logger.info(f"上传成功: {url}")
    return url


def upload_images(image_paths):
    """
    批量上传图片，返回 URL 列表
    
    Args:
        image_paths: 图片路径列表（本地路径或已有URL）

    Returns:
        URL 列表
    """
    urls = []
    for path in image_paths:
        if path.startswith('http://') or path.startswith('https://'):
            # 已经是 URL，直接使用
            urls.append(path)
        else:
            url = upload_image_to_cos(path)
            urls.append(url)
    return urls


# ============================================================
# 核心业务逻辑
# ============================================================
def create_goods(conn, goods_data, dry_run=False):
    """
    步骤1：创建商品主体（litemall_goods + litemall_goods_specification + litemall_goods_product）
    
    Args:
        conn: 数据库连接
        goods_data: dict, 包含以下字段：
            - name: 商品名称（必填）
            - pic_url: 主图URL（必填）
            - gallery: 轮播图URL列表（可选，默认=[pic_url]）
            - unit: 单位（默认=盆）
            - category_id: 分类ID（默认=0）
            - detail: 商品详情HTML（可选）
            - specifications: 规格值，如 "45" 或 "1盆"（必填）
            - spec_name: 规格名（默认="规格"）
            - sku_price: SKU价格（默认=0）
            - sku_url: SKU规格图片URL（可选，默认=pic_url）
        dry_run: 只打印SQL不执行

    Returns:
        (goods_id, product_id) 元组
    """
    name = goods_data['name']
    pic_url = goods_data['pic_url']
    gallery = goods_data.get('gallery', [pic_url])
    unit = goods_data.get('unit') or '盆'
    category_id = goods_data.get('category_id', DEFAULTS['category_id'])
    brand_id = goods_data.get('brand_id', DEFAULTS['brand_id'])
    detail = goods_data.get('detail', None)
    is_on_sale = goods_data.get('is_on_sale', DEFAULTS['is_on_sale'])
    sort_order = goods_data.get('sort_order', DEFAULTS['sort_order'])
    counter_price = goods_data.get('counter_price', DEFAULTS['counter_price'])
    retail_price = goods_data.get('retail_price', DEFAULTS['retail_price'])

    spec_value = goods_data['specifications']
    spec_name = goods_data.get('spec_name', '规格')
    sku_price = goods_data.get('sku_price', Decimal('0.00'))
    sku_url = goods_data.get('sku_url', pic_url)

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor = conn.cursor()

    # --- 检查同名商品是否已存在 ---
    cursor.execute(
        "SELECT id FROM litemall_goods WHERE name = %s AND deleted = 0 LIMIT 1",
        (name,)
    )
    existing = cursor.fetchone()

    if existing:
        goods_id = existing['id']
        if dry_run:
            logger.info(f"[DRY-RUN] 商品已存在: goods_id={goods_id}, name={name}，将追加规格 {spec_value}")
        else:
            logger.info(f"✓ 商品已存在: goods_id={goods_id}, name={name}，追加新规格，更新商品主体价格为 {retail_price}")
            cursor.execute(
                "UPDATE litemall_goods SET retail_price = %s, counter_price = %s WHERE id = %s",
                (retail_price, counter_price, goods_id)
            )
    else:
        # --- INSERT litemall_goods（新商品）---
        goods_sn = generate_goods_sn()
        gallery_json = json.dumps(gallery, ensure_ascii=False)

        if dry_run:
            logger.info(f"[DRY-RUN] INSERT litemall_goods: sn={goods_sn}, name={name}")
            return None, None

        sql_goods = """
            INSERT INTO litemall_goods (
                goods_sn, name, category_id, brand_id, gallery,
                pic_url, is_on_sale, sort_order, unit,
                counter_price, retail_price, detail,
                add_time, update_time, deleted
            ) VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, 0
            )
        """
        params_goods = (
            goods_sn, name, category_id, brand_id, gallery_json,
            pic_url, is_on_sale, sort_order, unit,
            counter_price, retail_price, detail,
            now, now
        )
        cursor.execute(sql_goods, params_goods)
        goods_id = cursor.lastrowid
        logger.info(f"✓ 创建商品: goods_id={goods_id}, sn={goods_sn}, name={name}")

    if dry_run:
        logger.info(f"[DRY-RUN] 追加规格: {spec_name}={spec_value}")
        return None, None

    # --- 检查同规格 SKU 是否已存在 ---
    spec_json = json.dumps([spec_value], ensure_ascii=False)
    cursor.execute(
        "SELECT id FROM litemall_goods_product WHERE goods_id = %s AND specifications = %s AND deleted = 0 LIMIT 1",
        (goods_id, spec_json)
    )
    existing_sku = cursor.fetchone()

    if existing_sku:
        product_id = existing_sku['id']
        logger.info(f"  ✓ SKU已存在: product_id={product_id}, spec={spec_value}，直接复用")
    else:
        # --- INSERT litemall_goods_specification ---
        spec_pic = goods_data.get('spec_pic_url', '')
        sql_spec = """
            INSERT INTO litemall_goods_specification (
                goods_id, specification, value, pic_url,
                add_time, update_time, deleted
            ) VALUES (%s, %s, %s, %s, %s, %s, 0)
        """
        cursor.execute(sql_spec, (goods_id, spec_name, spec_value, spec_pic, now, now))
        logger.info(f"  ✓ 创建规格: {spec_name}={spec_value}")

        # --- INSERT litemall_goods_product (SKU) ---
        sql_product = """
            INSERT INTO litemall_goods_product (
                goods_id, specifications, price, number, url,
                add_time, update_time, deleted
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, 0)
        """
        cursor.execute(sql_product, (goods_id, spec_json, sku_price, 0, sku_url, now, now))
        product_id = cursor.lastrowid
        logger.info(f"  ✓ 创建SKU: product_id={product_id}, spec={spec_value}, price={sku_price}")

    return goods_id, product_id


def create_stock_batch(conn, batch_data, dry_run=False):
    """
    步骤2：创建库存批次（litemall_stock_batch + litemall_stock_batch_region + litemall_stock_log）
    按照文档第三章逻辑。

    Args:
        conn: 数据库连接
        batch_data: dict, 包含以下字段：
            - product_id: SKU ID（必填）
            - goods_id: 商品ID（必填）
            - creator_id: 团长用户ID（必填）
            - total_quantity: 入库数量（必填，>0）
            - cost_price / cost_total: 采购价（至少一个）
            - wholesale_price / wholesale_total: 批发价（至少一个）
            - scope: PRIVATE/PUBLIC（默认PRIVATE）
            - remark: 备注（可选）
            - regions: 城市列表 [{'code': 320100, 'name': '南京市'}, ...]（必填，至少1个）
        dry_run: 只打印SQL不执行

    Returns:
        batch_id
    """
    product_id = batch_data['product_id']
    goods_id = batch_data['goods_id']
    creator_id = batch_data['creator_id']
    total_quantity = batch_data['total_quantity']
    scope = batch_data.get('scope') or 'PUBLIC'
    remark = batch_data.get('remark', None)
    regions = batch_data['regions']

    if total_quantity <= 0:
        raise ValueError(f"total_quantity 必须 > 0，当前值: {total_quantity}")
    if not regions:
        raise ValueError("regions 不能为空，至少需要 1 个城市")

    # 价格换算
    cost_price, cost_total, wholesale_price, wholesale_total = calculate_prices(
        cost_price=batch_data.get('cost_price'),
        cost_total=batch_data.get('cost_total'),
        wholesale_price=batch_data.get('wholesale_price'),
        wholesale_total=batch_data.get('wholesale_total'),
        quantity=total_quantity
    )

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor = conn.cursor()

    if dry_run:
        logger.info(f"[DRY-RUN] INSERT litemall_stock_batch: product_id={product_id}, qty={total_quantity}")
        logger.info(f"[DRY-RUN] INSERT litemall_stock_batch_region: {len(regions)} 个城市")
        logger.info(f"[DRY-RUN] INSERT litemall_stock_log: type=INBOUND")
        return None

    # --- 步骤1: INSERT litemall_stock_batch ---
    sql_batch = """
        INSERT INTO litemall_stock_batch (
            product_id, goods_id, creator_id, scope,
            cost_price, cost_total, wholesale_price, wholesale_total,
            total_quantity, sold_quantity, adjust_quantity,
            remark, status, deleted, add_time, update_time
        ) VALUES (
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, 0, 0,
            %s, 'AVAILABLE', 0, %s, %s
        )
    """
    cursor.execute(sql_batch, (
        product_id, goods_id, creator_id, scope,
        cost_price, cost_total, wholesale_price, wholesale_total,
        total_quantity,
        remark, now, now
    ))
    batch_id = cursor.lastrowid
    logger.info(f"  ✓ 创建批次: batch_id={batch_id}, qty={total_quantity}, "
                f"cost={cost_price}/{cost_total}, wholesale={wholesale_price}/{wholesale_total}")

    # --- 步骤2: INSERT litemall_stock_batch_region ---
    sql_region = """
        INSERT INTO litemall_stock_batch_region (
            batch_id, region_code, region_name, add_time, deleted
        ) VALUES (%s, %s, %s, %s, 0)
    """
    for region in regions:
        cursor.execute(sql_region, (batch_id, region['code'], region.get('name', ''), now))
        logger.info(f"    ✓ 绑定城市: {region.get('name', '')}({region['code']})")

    # --- 步骤3: INSERT litemall_stock_log ---
    sql_log = """
        INSERT INTO litemall_stock_log (
            batch_id, type, quantity, before_remaining, after_remaining,
            order_id, activity_id, reason, operator_id, add_time
        ) VALUES (
            %s, 'INBOUND', %s, 0, %s,
            NULL, NULL, '创建批次入库', %s, %s
        )
    """
    cursor.execute(sql_log, (batch_id, total_quantity, total_quantity, creator_id, now))
    logger.info(f"    ✓ 入库日志: INBOUND, 0 → {total_quantity}")

    return batch_id


def adjust_stock(conn, adjust_data, dry_run=False):
    """
    调整库存（文档第四章逻辑）

    Args:
        conn: 数据库连接
        adjust_data: dict, 包含：
            - batch_id: 批次ID（必填）
            - adjust_type: ADD/SUB（必填）
            - quantity: 调整数量，正整数（必填）
            - reason: 原因（可选）
            - operator_id: 操作人ID（必填）
        dry_run: 只打印SQL不执行

    Returns:
        after_remaining
    """
    batch_id = adjust_data['batch_id']
    adjust_type = adjust_data['adjust_type'].upper()
    quantity = adjust_data['quantity']
    reason = adjust_data.get('reason', None)
    operator_id = adjust_data['operator_id']

    if adjust_type not in ('ADD', 'SUB'):
        raise ValueError(f"adjust_type 必须是 ADD 或 SUB，当前: {adjust_type}")
    if quantity <= 0:
        raise ValueError(f"quantity 必须 > 0，当前: {quantity}")

    cursor = conn.cursor()

    # 读取当前批次
    cursor.execute("""
        SELECT total_quantity, adjust_quantity, sold_quantity, status
        FROM litemall_stock_batch
        WHERE id = %s AND deleted = 0
        FOR UPDATE
    """, (batch_id,))
    batch = cursor.fetchone()
    if not batch:
        raise ValueError(f"批次 {batch_id} 不存在或已删除")

    tq = batch['total_quantity']
    aq = batch['adjust_quantity']
    sq = batch['sold_quantity']
    old_status = batch['status']
    before_remaining = tq + aq - sq

    if adjust_type == 'ADD':
        after_remaining = before_remaining + quantity
        new_aq = aq + quantity
        log_type = 'ADJUST_ADD'
    else:  # SUB
        if quantity > before_remaining:
            raise ValueError(f"调减数量({quantity})超过可用库存({before_remaining})！")
        after_remaining = before_remaining - quantity
        new_aq = aq - quantity
        log_type = 'ADJUST_SUB'

    # status 更新规则
    if after_remaining == 0:
        new_status = 'EXHAUSTED'
    elif after_remaining > 0 and old_status == 'EXHAUSTED':
        new_status = 'AVAILABLE'
    else:
        new_status = old_status

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if dry_run:
        logger.info(f"[DRY-RUN] ADJUST batch_id={batch_id}: {adjust_type} {quantity}, "
                     f"remaining: {before_remaining} → {after_remaining}")
        return after_remaining

    # UPDATE 批次
    cursor.execute("""
        UPDATE litemall_stock_batch
        SET adjust_quantity = %s, status = %s, update_time = %s
        WHERE id = %s AND deleted = 0
    """, (new_aq, new_status, now, batch_id))

    # INSERT 日志
    cursor.execute("""
        INSERT INTO litemall_stock_log (
            batch_id, type, quantity, before_remaining, after_remaining,
            order_id, activity_id, reason, operator_id, add_time
        ) VALUES (%s, %s, %s, %s, %s, NULL, NULL, %s, %s, %s)
    """, (batch_id, log_type, quantity, before_remaining, after_remaining,
          reason, operator_id, now))

    logger.info(f"  ✓ 调整完成: batch_id={batch_id}, {log_type} {quantity}, "
                f"remaining: {before_remaining} → {after_remaining}, status: {new_status}")

    return after_remaining


# ============================================================
# 完整上架流程（商品创建 + 库存创建）
# ============================================================
def full_listing(product_data, dry_run=False):
    """
    完整上架一个商品：创建商品 → 创建SKU → 创建库存批次

    Args:
        product_data: dict, 包含所有必要信息：
            商品信息：
                - name: 商品名称
                - pic_url: 主图URL（或本地路径，会自动上传COS）
                - gallery: 轮播图列表（可选）
                - unit: 单位（默认=盆）
                - specifications: 规格值
                - detail: 详情HTML（可选）
            库存信息：
                - creator_id: 团长用户ID
                - total_quantity: 入库数量
                - cost_price: 采购单价
                - wholesale_price: 批发单价
                - regions: 城市列表 [{'code': 320100, 'name': '南京市'}]
                - remark: 备注（可选）
        dry_run: True=只打印不执行

    Returns:
        dict: {goods_id, product_id, batch_id}
    """
    conn = get_connection()
    try:
        conn.begin()
        cursor = conn.cursor()

        # 预先检查商品规格 (SKU) 是否已存在于数据库
        name = product_data['name']
        spec_value = product_data['specifications']
        spec_json = json.dumps([spec_value], ensure_ascii=False)

        cursor.execute(
            "SELECT id FROM litemall_goods WHERE name = %s AND deleted = 0 LIMIT 1",
            (name,)
        )
        existing_goods = cursor.fetchone()

        existing_sku = None
        if existing_goods:
            goods_id = existing_goods['id']
            cursor.execute(
                "SELECT id FROM litemall_goods_product WHERE goods_id = %s AND specifications = %s AND deleted = 0 LIMIT 1",
                (goods_id, spec_json)
            )
            existing_sku = cursor.fetchone()

        if existing_goods and existing_sku:
            # 商品与规格均已存在，跳过图片上传与创建
            goods_id = existing_goods['id']
            product_id = existing_sku['id']
            logger.info(f"  ✓ 商品及规格已存在于数据库: goods_id={goods_id}, product_id={product_id}，跳过图片上传与商品创建")
            
            if not dry_run:
                selling_price = product_data.get('wholesale_price') or Decimal('0.00')
                cursor.execute(
                    "UPDATE litemall_goods SET retail_price = %s, counter_price = %s WHERE id = %s",
                    (selling_price, selling_price, goods_id)
                )
        else:
            # 需要新增商品（goods不存在）或新增规格（goods存在但sku不存在）
            pic_url = product_data['pic_url']
            gallery = product_data.get('gallery', [pic_url])

            # 判断是否仅仅需要新增规格（商品已存在，只需新增规格）
            only_add_spec = (existing_goods is not None)

            if only_add_spec:
                # 仅新增规格时：只上传主图/规格图(pic_url)，不上传轮播图(gallery)
                logger.info(f"  ℹ️ 商品已存在，仅新增规格。只上传规格图，跳过轮播图上传。")
                if dry_run:
                    if not pic_url.startswith('http'):
                        logger.info(f"  [DRY-RUN] 将上传主图/规格图: {pic_url}")
                    gallery = []
                else:
                    if not pic_url.startswith('http'):
                        pic_url = upload_image_to_cos(pic_url)
                    gallery = []
            else:
                # 崭新商品：必须上传主图和所有轮播图
                if dry_run:
                    # DRY-RUN：不上传，仅显示将要处理的图片
                    if not pic_url.startswith('http'):
                        logger.info(f"  [DRY-RUN] 将上传主图: {pic_url}")
                    for g in gallery:
                        if not g.startswith('http'):
                            logger.info(f"  [DRY-RUN] 将上传轮播图: {g}")
                else:
                    # 正式执行：上传图片，使用缓存避免同一文件重复上传
                    uploaded_cache = {}  # local_path -> cos_url
                    if not pic_url.startswith('http'):
                        pic_url = upload_image_to_cos(pic_url)
                        uploaded_cache[product_data['pic_url']] = pic_url

                    uploaded_gallery = []
                    for g in gallery:
                        if g.startswith('http'):
                            uploaded_gallery.append(g)
                        elif g in uploaded_cache:
                            uploaded_gallery.append(uploaded_cache[g])
                            logger.info(f"  复用已上传图片: {g} -> {uploaded_cache[g]}")
                        else:
                            url = upload_image_to_cos(g)
                            uploaded_cache[g] = url
                            uploaded_gallery.append(url)
                    gallery = uploaded_gallery

            # 步骤1: 创建商品
            selling_price = product_data.get('wholesale_price') or Decimal('0.00')
            goods_data = {
                'name': product_data['name'],
                'pic_url': pic_url,
                'gallery': gallery,
                'unit': product_data.get('unit') or '盆',
                'category_id': product_data.get('category_id', DEFAULTS['category_id']),
                'specifications': product_data['specifications'],
                'spec_name': product_data.get('spec_name', '规格'),
                'sku_price': product_data.get('sku_price') or selling_price,
                'sku_url': product_data.get('sku_url', pic_url),
                'detail': product_data.get('detail', None),
                'retail_price': selling_price,
                'counter_price': selling_price,
            }

            goods_id, product_id = create_goods(conn, goods_data, dry_run=dry_run)

        if dry_run:
            logger.info(f"[DRY-RUN] 完整上架流程预览完毕: {product_data['name']}")
            conn.rollback()
            return {'goods_id': None, 'product_id': None, 'batch_id': None}

        # 步骤2: 创建库存批次
        batch_data = {
            'product_id': product_id,
            'goods_id': goods_id,
            'creator_id': product_data['creator_id'],
            'total_quantity': product_data['total_quantity'],
            'cost_price': product_data.get('cost_price'),
            'cost_total': product_data.get('cost_total'),
            'wholesale_price': product_data.get('wholesale_price'),
            'wholesale_total': product_data.get('wholesale_total'),
            'scope': product_data.get('scope') or 'PUBLIC',
            'remark': product_data.get('remark'),
            'regions': product_data['regions'],
        }

        batch_id = create_stock_batch(conn, batch_data, dry_run=dry_run)

        conn.commit()
        logger.info(f"★ 上架完成: name={product_data['name']}, "
                     f"goods_id={goods_id}, product_id={product_id}, batch_id={batch_id}")

        return {
            'goods_id': goods_id,
            'product_id': product_id,
            'batch_id': batch_id,
        }

    except Exception as e:
        conn.rollback()
        logger.error(f"✗ 上架失败，已回滚: {e}")
        raise
    finally:
        conn.close()


def batch_listing(products_list, dry_run=False):
    """
    批量上架商品

    Args:
        products_list: 商品数据列表
        dry_run: 只打印不执行

    Returns:
        批量结果统计
    """
    results = {
        'total': len(products_list),
        'success': 0,
        'failed': 0,
        'details': [],
        'errors': [],
    }

    for i, product in enumerate(products_list, 1):
        logger.info(f"\n{'='*60}")
        logger.info(f"正在上架第 {i}/{len(products_list)} 个商品: {product.get('name', '未知')}")
        logger.info(f"{'='*60}")

        try:
            result = full_listing(product, dry_run=dry_run)
            results['success'] += 1
            results['details'].append({
                'index': i,
                'name': product.get('name'),
                **result
            })
        except Exception as e:
            results['failed'] += 1
            results['errors'].append({
                'index': i,
                'name': product.get('name'),
                'error': str(e)
            })
            logger.error(f"第 {i} 个商品上架失败: {e}")

    logger.info(f"\n{'='*60}")
    logger.info(f"批量上架完成: 总计 {results['total']}, "
                f"成功 {results['success']}, 失败 {results['failed']}")
    if results['errors']:
        logger.warning("失败列表:")
        for err in results['errors']:
            logger.warning(f"  #{err['index']} {err['name']}: {err['error']}")
    logger.info(f"{'='*60}")

    return results


# ============================================================
# Excel 数据读取
# ============================================================
def load_products_from_excel(excel_path):
    """
    从 Excel 模板读取商品数据，返回 product_data 列表
    """
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['商品上架数据']

    products = []
    # 数据从第3行开始（第1行=表头，第2行=说明）
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name or str(name).strip() == '':
            continue  # 跳过空行

        pic_url = str(ws.cell(row=row, column=2).value or '').strip()
        gallery_raw = str(ws.cell(row=row, column=3).value or '').strip()
        spec_value = str(ws.cell(row=row, column=4).value or '').strip()
        unit = str(ws.cell(row=row, column=5).value or '').strip() or '盆'
        creator_name = str(ws.cell(row=row, column=6).value or '').strip()
        creator_id = resolve_creator_id(creator_name)
        total_quantity = int(ws.cell(row=row, column=7).value or 0)
        cost_price = ws.cell(row=row, column=8).value
        wholesale_price = ws.cell(row=row, column=9).value

        # 合并所有城市输入源：
        # J列(10)=省份快捷, K-O列(11-15)=城市1-5, P列(16)=手动输入
        city_parts = []
        
        # 1. 快捷选择
        shortcut = str(ws.cell(row=row, column=10).value or '').strip()
        if shortcut:
            city_parts.append(shortcut)
            
        # 2. 城市1-城市5
        for col_offset in range(11, 16):
            c_val = str(ws.cell(row=row, column=col_offset).value or '').strip()
            if c_val:
                city_parts.append(c_val)
                
        # 3. 手动输入城市
        manual_city = str(ws.cell(row=row, column=16).value or '').strip()
        if manual_city:
            city_parts.append(manual_city)

        # 用 | 连接成最终字符串，resolve_city_regions 会自动去重与匹配
        city_names_raw = '|'.join(city_parts)
        
        # Q列(17)=备注
        remark = ws.cell(row=row, column=17).value

        # 解析图片路径（支持本地路径自动修复）
        if pic_url and not pic_url.startswith('http'):
            pic_url = resolve_image_path(pic_url)

        # 解析轮播图
        gallery = []
        if gallery_raw:
            for g in gallery_raw.split('|'):
                g = g.strip()
                if g:
                    if not g.startswith('http'):
                        g = resolve_image_path(g)
                    gallery.append(g)
        if not gallery:
            gallery = [pic_url]

        # 解析城市（通过名称自动匹配编码）
        regions = resolve_city_regions(city_names_raw)

        product = {
            'name': str(name).strip(),
            'pic_url': pic_url,
            'gallery': gallery,
            'specifications': spec_value,
            'unit': unit,
            'creator_id': creator_id,
            'total_quantity': total_quantity,
            'cost_price': Decimal(str(cost_price)) if cost_price else None,
            'wholesale_price': Decimal(str(wholesale_price)) if wholesale_price else None,
            'regions': regions,
            'remark': str(remark) if remark else None,
        }
        products.append(product)
        logger.info(f"  读取第{row}行: {name}, 规格={spec_value}, 数量={total_quantity}, "
                    f"采购价={cost_price}, 批发价={wholesale_price}, 城市={regions}")

    wb.close()
    logger.info(f"共读取 {len(products)} 条商品数据")
    return products


def process_pipeline_excel(excel_path, dry_run=False):
    """
    解析并执行管线模式的商品上架
    """
    if not dry_run:
        # 检查 Excel 文件是否被锁定（若被 Excel/WPS 等软件打开，循环提示用户手动关闭并重试）
        while True:
            try:
                with open(excel_path, 'r+b') as f:
                    pass
                break # 未被锁定，正常跳出
            except IOError:
                logger.warning(f"\n⚠️  检测到 Excel 文件被锁定或被占用，请先关闭 Excel/WPS 软件！")
                logger.warning(f"文件路径: {excel_path}")
                choice = input("请在关闭表格软件后，按回车键 [Enter] 重试，或输入 q 退出上架: ").strip().lower()
                if choice == 'q':
                    logger.info("已手动取消上架。")
                    import sys
                    sys.exit(0)


    wb = openpyxl.load_workbook(excel_path)
    ws = wb['商品上架数据']

    pending_dir = os.path.dirname(excel_path)
    completed_dir = os.path.join(os.path.dirname(pending_dir), 'completed_upload')

    if not dry_run:
        os.makedirs(completed_dir, exist_ok=True)

    rows_to_delete = []
    failed_count = 0
    total_processed = 0

    results = {
        'total': 0,
        'success': 0,
        'failed': 0,
        'errors': []
    }

    # 从第3行开始
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        img_dir = ws.cell(row=r, column=17).value
        # 如果商品名称和图片目录都为空，视为结束或空行
        if (not name or str(name).strip() == '') and (not img_dir or str(img_dir).strip() == ''):
            continue

        total_processed += 1
        name = str(name or '未定义商品').strip()
        logger.info(f"\n{'='*60}")
        logger.info(f"正在处理第 {r} 行: {name} (图片目录={img_dir})")
        logger.info(f"{'='*60}")

        try:
            # 基础验证
            if not img_dir or str(img_dir).strip() == '':
                raise ValueError("图片目录未填写")
            
            img_dir_name = str(img_dir).strip()
            img_dir_path = os.path.join(pending_dir, img_dir_name)
            if not os.path.isdir(img_dir_path):
                raise FileNotFoundError(f"找不到指定的图片目录: pending_upload/{img_dir_name}")

            # 动态图源匹配
            patterns = ["*.png", "*.jpg", "*.jpeg", "*.PNG", "*.JPG"]
            img_files = []
            for p in patterns:
                img_files.extend(glob.glob(os.path.join(img_dir_path, p)))

            if not img_files:
                raise FileNotFoundError(f"目录 pending_upload/{img_dir_name} 下未找到任何图片文件")

            # 自然排序
            img_files.sort(key=lambda x: os.path.basename(x))
            pic_url = img_files[0]
            gallery = img_files[1:] if len(img_files) > 1 else [pic_url]

            # 提取其它数据
            spec_value = str(ws.cell(row=r, column=4).value or '').strip()
            if not spec_value:
                raise ValueError("规格值不能为空")

            unit = str(ws.cell(row=r, column=5).value or '').strip() or '盆'
            
            creator_name = str(ws.cell(row=r, column=6).value or '').strip()
            if not creator_name:
                raise ValueError("团长名称不能为空")
            creator_id = resolve_creator_id(creator_name)

            total_qty_val = ws.cell(row=r, column=7).value
            if total_qty_val is None:
                raise ValueError("入库数量不能为空")
            total_quantity = int(total_qty_val)
            if total_quantity <= 0:
                raise ValueError(f"入库数量必须 > 0，当前为 {total_quantity}")

            cost_price = ws.cell(row=r, column=8).value
            wholesale_price = ws.cell(row=r, column=9).value

            # 合并省市
            city_parts = []
            shortcut = str(ws.cell(row=r, column=10).value or '').strip()
            if shortcut:
                city_parts.append(shortcut)
            for col_offset in range(11, 16):
                c_val = str(ws.cell(row=r, column=col_offset).value or '').strip()
                if c_val:
                    city_parts.append(c_val)
            manual_city = str(ws.cell(row=r, column=16).value or '').strip()
            if manual_city:
                city_parts.append(manual_city)

            city_names_raw = '|'.join(city_parts)
            regions = resolve_city_regions(city_names_raw)

            remark = ws.cell(row=r, column=19).value

            # 公共/私有
            scope_raw = str(ws.cell(row=r, column=18).value or '').strip()
            scope_map = {'公共': 'PUBLIC', '私有': 'PRIVATE'}
            scope = scope_map.get(scope_raw, 'PUBLIC')

            product_data = {
                'name': name,
                'pic_url': pic_url,
                'gallery': gallery,
                'specifications': spec_value,
                'unit': unit,
                'creator_id': creator_id,
                'total_quantity': total_quantity,
                'cost_price': Decimal(str(cost_price)) if cost_price else None,
                'wholesale_price': Decimal(str(wholesale_price)) if wholesale_price else None,
                'regions': regions,
                'remark': str(remark) if remark else None,
                'scope': scope,
            }

            # 执行上架
            result = full_listing(product_data, dry_run=dry_run)

            if not dry_run:
                rows_to_delete.append(r)
                # 移动图片目录到已完成归档箱
                dest_dir_path = os.path.join(completed_dir, img_dir_name)
                # 如果已存在目标目录，加上时间戳避免冲突
                if os.path.exists(dest_dir_path):
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    dest_dir_path = os.path.join(completed_dir, f"{img_dir_name}_{ts}")
                shutil.move(img_dir_path, dest_dir_path)
                logger.info(f"  ✓ 归档图片目录: {img_dir_name} -> completed_upload/")
            else:
                logger.info(f"[DRY-RUN] 行 {r} 校验并模拟上架成功: {name}")

        except Exception as e:
            failed_count += 1
            error_msg = str(e)
            logger.error(f"行 {r} 上架失败: {error_msg}")
            results['errors'].append({'row': r, 'name': name, 'error': error_msg})

            if not dry_run:
                # 回写错误信息到第20列
                ws.cell(row=r, column=20, value=error_msg)

    results['total'] = total_processed
    results['success'] = (total_processed - failed_count)
    results['failed'] = failed_count

    if not dry_run:
        # 从后往前删除成功的行
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r)
        wb.save(excel_path)
        logger.info("\n📝 成功从 Excel 中删除已完成行，并保存回写错误信息")

    wb.close()
    return results


def create_empty_pipeline_excel(excel_path):
    """
    创建一个结构和数据验证完全一致的空 pipeline Excel
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品上架数据"

    # Styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    opt_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
    helper_fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type="solid")
    desc_font = Font(name="微软雅黑", size=9, color="888888", italic=True)

    columns = [
        ("商品名称",       20, True,  "如：散尾葵、龟背竹"),
        ("主图\n(已弃用)",   20, False, "留空即可，系统自动扫描图片目录"),
        ("轮播图\n(已弃用)", 25, False, "留空即可，系统自动扫描图片目录"),
        ("规格值",         12, True,  "如：45、1盆、大号"),
        ("单位",           8,  False, "默认：盆"),
        ("团长名称",       12, True,  "森屿家用户，如李琳"),
        ("入库数量",       10, True,  "必须 > 0"),
        ("采购单价",       10, True,  "元"),
        ("批发单价",       10, True,  "元"),
        ("省份快捷选择",   15, False, "下拉选择省份（整省全部上架）"),
        ("城市1",          12, False, "下拉选择城市"),
        ("城市2",          12, False, "下拉选择城市"),
        ("城市3",          12, False, "下拉选择城市"),
        ("城市4",          12, False, "下拉选择城市"),
        ("城市5",          12, False, "下拉选择城市"),
        ("手动输入城市",   20, False, "若超5个城市或需手填，用 | 分隔"),
        ("图片目录",       25, True,  "系统自动填充的图片目录"),
        ("公共/私有",      10, False, "下拉选择：公共 或 私有"),
        ("备注",           20, False, "可选"),
        ("错误信息\n(系统回写)", 30, False, "上架失败时系统自动回写错误原因"),
    ]

    for col_idx, (name, width, req, desc) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=f"{name}\n{'(必填)' if req else '(选填)'}")
        cell.font = header_font
        if "城市" in name or "省份" in name:
            cell.fill = helper_fill
        else:
            cell.fill = header_fill if req else opt_fill
        cell.alignment = wrap_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 40

    for col_idx, (_, _, _, desc) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font = desc_font
        cell.alignment = wrap_alignment
        cell.border = thin_border

    # 下拉源数据 sheet
    ws_list = wb.create_sheet("下拉源数据")
    ws_list.cell(row=1, column=1, value="省份快捷")
    shortcut_options = ["江苏省全部", "浙江省全部", "湖北省全部", "广东省全部", "上海市"]
    for r_idx, s in enumerate(shortcut_options, 2):
        ws_list.cell(row=r_idx, column=1, value=s)

    # 城市列表数据
    city_data = {
        "江苏省": ["南京市", "无锡市", "徐州市", "常州市", "苏州市", "南通市", "连云港市", "淮安市", "盐城市", "扬州市", "镇江市", "泰州市", "宿迁市"],
        "浙江省": ["杭州市", "宁波市", "温州市", "嘉兴市", "湖州市", "绍兴市", "金华市", "衢州市", "舟山市", "台州市", "丽水市"],
        "湖北省": ["武汉市", "黄石市", "十堰市", "宜昌市", "襄阳市", "易用市", "荆门市", "孝感市", "荆州市", "黄冈市", "咸宁市", "随州市", "恩施州"],
        "广东省": ["广州市", "韶关市", "深圳市", "珠海市", "汕头市", "佛山市", "江门市", "湛江市", "茂名市", "肇庆市", "惠州市", "梅州市", "汕尾市", "河源市", "阳江市", "清远市", "东莞市", "中山市", "潮州市", "揭阳市", "云浮市"],
        "上海市": ["上海市"],
    }
    all_cities = []
    for prov, cities in city_data.items():
        for city_name in cities:
            if city_name not in all_cities:
                all_cities.append(city_name)

    ws_list.cell(row=1, column=2, value="所有城市")
    for r_idx, c in enumerate(all_cities, 2):
        ws_list.cell(row=r_idx, column=2, value=c)

    # 团长列表数据
    creators = get_all_creators()
    ws_list.cell(row=1, column=3, value="所有团长")
    for r_idx, name in enumerate(creators, 2):
        ws_list.cell(row=r_idx, column=3, value=name)

    ws_list.views.sheetView[0].showGridLines = True

    # 数据验证
    dv_shortcut = DataValidation(
        type="list",
        formula1=f"=下拉源数据!$A$2:$A${len(shortcut_options) + 1}",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="选项无效",
        error="请从下拉列表中选择省份快捷选项",
    )
    ws.add_data_validation(dv_shortcut)
    dv_shortcut.add("J3:J200")

    dv_cities = DataValidation(
        type="list",
        formula1=f"=下拉源数据!$B$2:$B${len(all_cities) + 1}",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="选项无效",
        error="请从下拉列表中选择支持的城市名称",
    )
    ws.add_data_validation(dv_cities)
    for col_idx in range(11, 16):
        col_letter = get_column_letter(col_idx)
        dv_cities.add(f"{col_letter}3:{col_letter}200")

    dv_creator = DataValidation(
        type="list",
        formula1=f"=下拉源数据!$C$2:$C${len(creators) + 1}",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="选项无效",
        error="请选择有效的团长名称",
    )
    ws.add_data_validation(dv_creator)
    dv_creator.add("F3:F200")

    dv_scope = DataValidation(
        type="list",
        formula1='"公共,私有"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="选项无效",
        error="请选择 公共 或 私有",
    )
    ws.add_data_validation(dv_scope)
    dv_scope.add("R3:R200")

    # 预填默认值到第3行
    ws.cell(row=3, column=5, value="盆")
    ws.cell(row=3, column=18, value="公共")

    wb.save(excel_path)
    wb.close()


def append_pipeline_excel_record(excel_path, plant_name, spec_value, dir_name):
    """
    往商品上架数据.xlsx 追加一行待上架记录
    """
    import openpyxl
    from openpyxl.styles import Font, Border, Side, Alignment
    
    if not os.path.exists(excel_path):
        logger.info(f"  Excel 文件不存在，正在初始化生成模板: pending_upload/商品上架数据.xlsx")
        create_empty_pipeline_excel(excel_path)
        
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["商品上架数据"]
    
    # 查找第一行空行（从第3行开始）
    target_row = 3
    while True:
        val = ws.cell(row=target_row, column=1).value
        # 如果第一列 (商品名称) 和第十七列 (图片目录) 都为空，则认为是空行
        if val is None and ws.cell(row=target_row, column=17).value is None:
            break
        target_row += 1
        
    # 追加新行
    ws.cell(row=target_row, column=1, value=plant_name)
    ws.cell(row=target_row, column=4, value=spec_value)
    ws.cell(row=target_row, column=5, value="盆")       # 单位默认：盆
    ws.cell(row=target_row, column=17, value=dir_name)
    ws.cell(row=target_row, column=18, value="公共")    # 公共/私有默认：公共
    
    # 设置样式
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    data_font = Font(name="微软雅黑", size=10)
    for col_idx in range(1, 21):
        cell = ws.cell(row=target_row, column=col_idx)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
    wb.save(excel_path)
    wb.close()
    logger.info(f"  ✓ 已在 Excel 第 {target_row} 行追加记录: 商品={plant_name}, 规格={spec_value}, 图片目录={dir_name}")


def check_batch_already_exists(b_id, pending_upload_dir, completed_dir):
    """
    检查指定批次 ID 是否已存在于 pending_upload 或 completed_upload 目录中
    返回匹配到的文件夹相对路径列表
    """
    matches = []
    
    dirs_to_scan = [
        ('pending_upload', pending_upload_dir),
        ('completed_upload', completed_dir)
    ]
    
    for label, base_path in dirs_to_scan:
        if not os.path.exists(base_path):
            continue
        for item in os.listdir(base_path):
            item_path = os.path.join(base_path, item)
            if not os.path.isdir(item_path):
                continue
                
            # 1. 检查文件夹前缀是否为 "批次号-"
            if item.startswith(f"{b_id}-"):
                matches.append(f"{label}/{item} (基于批次号前缀匹配)")
                continue
                
            # 2. 精确检查: 是否有 batch_info.json 且匹配
            info_file = os.path.join(item_path, 'batch_info.json')
            if os.path.exists(info_file):
                try:
                    with open(info_file, 'r', encoding='utf-8') as jf:
                        info = json.load(jf)
                        if str(info.get('batch_id')) == str(b_id):
                            matches.append(f"{label}/{item} (基于 batch_info.json 匹配)")
                            continue
                except Exception:
                    pass
                
    return list(set(matches))


def pull_ai_generated_data():
    """
    拉取 AI 生图批次数据到 pending_upload 目录中，并追加/更新 Excel 记录
    """
    import re
    if not os.path.exists(AI_OUTPUT_DIR):
        logger.error(f"AI 生图输出目录不存在: {AI_OUTPUT_DIR}")
        input("\n按回车键 [Enter] 返回主菜单...")
        return

    # 1. 扫描 AI 生图输出目录中的所有子文件夹
    subdirs = []
    for item in os.listdir(AI_OUTPUT_DIR):
        full_path = os.path.join(AI_OUTPUT_DIR, item)
        if os.path.isdir(full_path):
            # 提取批次号 (前缀数字)
            match = re.match(r'^(\d+)-', item)
            if match:
                batch_id = int(match.group(1))
                subdirs.append({
                    'name': item,
                    'path': full_path,
                    'batch_id': batch_id
                })

    if not subdirs:
        logger.warning(f"在 {AI_OUTPUT_DIR} 下没有找到形如 '数字-' 前缀的批次文件夹！")
        input("\n按回车键 [Enter] 返回主菜单...")
        return

    # 按批次号分组并排序
    batch_groups = {}
    for d in subdirs:
        b_key = str(d['batch_id'])
        if b_key not in batch_groups:
            batch_groups[b_key] = []
        batch_groups[b_key].append(d)

    sorted_batches = sorted([int(k) for k in batch_groups.keys()])

    while True:
        logger.info(f"\n{'='*60}")
        logger.info("可用的 AI 生图批次列表：")
        logger.info(f"{'='*60}")
        for b_id in sorted_batches:
            folders = batch_groups[str(b_id)]
            # 尝试从每个文件夹下解析植物名 and 规格
            desc_list = []
            for f in folders:
                meta_path = os.path.join(str(f['path']), 'batch_meta.json')
                plant_name = "未知植物"
                height = "未定义规格"
                if os.path.exists(meta_path):
                    try:
                        with open(meta_path, 'r', encoding='utf-8') as jf:
                            meta = json.load(jf)
                            plant_name = meta.get('plant_name', plant_name)
                            height = meta.get('height', height)
                    except Exception:
                        pass
                desc_list.append(f"{str(f['name'])} (植物: {plant_name}, 规格: {height})")
            
            logger.info(f"  批次 [{b_id}]:")
            for desc in desc_list:
                logger.info(f"    - {desc}")

        logger.info(f"{'='*60}")
        choice = input("\n请输入要拉取的批次 (如 1 或 1-3, 输入 q 返回): ").strip()
        if choice.lower() == 'q':
            logger.info("已返回主菜单。")
            return

        selected_batch_ids = []
        if '-' in choice:
            parts = choice.split('-', 1)
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                start, end = int(parts[0]), int(parts[1])
                selected_batch_ids = list(range(start, end + 1))
        elif choice.isdigit():
            selected_batch_ids = [int(choice)]

        if not selected_batch_ids:
            logger.error(f"输入格式无效: {choice}")
            input("\n按回车键 [Enter] 返回...")
            continue

        # 检查选择的批次是否全部有效
        valid_batches = [b for b in selected_batch_ids if str(b) in batch_groups]
        if not valid_batches:
            logger.error("所选批次均无效或不存在！")
            input("\n按回车键 [Enter] 返回...")
            continue

        logger.info(f"将要拉取批次: {valid_batches}")

        # 目标上架暂存目录与已完成归档目录
        base_dir = os.path.dirname(os.path.abspath(__file__))
        pending_upload_dir = os.path.join(base_dir, 'pending_upload')
        completed_dir = os.path.join(base_dir, 'completed_upload')
        os.makedirs(pending_upload_dir, exist_ok=True)
        excel_path = os.path.join(pending_upload_dir, '商品上架数据.xlsx')

        cancelled_by_user = False

        # 遍历并处理
        for b_id in valid_batches:
            if cancelled_by_user:
                break

            # ==========================================================
            # 查重逻辑：在批次级别检查整个批次号是否已存在
            # ==========================================================
            existing_matches = check_batch_already_exists(b_id, pending_upload_dir, completed_dir)
            if existing_matches:
                logger.warning(f"\n⚠️  警告：检测到批次 [{b_id}] 已经运行/拉取过！")
                logger.warning(f"   已存在于：")
                for match_path in existing_matches:
                    logger.warning(f"     - {match_path}")
                user_confirm = input("   是否仍要重复执行拉取该批次？(y/n) [直接回车或输入 y 继续, 输入 n 返回上一步]: ").strip().lower()
                if user_confirm == 'n':
                    logger.info(f"已取消拉取批次 {b_id}，返回批次选择。")
                    cancelled_by_user = True
                    break
                else:
                    logger.info(f"坚持执行拉取批次 {b_id} 的所有商品。")

            folders = batch_groups[str(b_id)]
            for f in folders:
                scenes_src_dir = os.path.join(str(f['path']), 'scenes')
                meta_path = os.path.join(str(f['path']), 'batch_meta.json')

                plant_name = str(f['name'])
                height = "60cm"

                # 解析 batch_meta.json
                if os.path.exists(meta_path):
                    try:
                        with open(meta_path, 'r', encoding='utf-8') as jf:
                            meta = json.load(jf)
                            plant_name = meta.get('plant_name', plant_name)
                            height = meta.get('height', height)
                    except Exception as e:
                        logger.warning(f"读取 batch_meta.json 失败: {e}")

                logger.info(f"\n正在拉取目录: {str(f['name'])}")

                # 检查 scenes 目录
                if not os.path.exists(scenes_src_dir) or not os.path.isdir(scenes_src_dir):
                    logger.warning(f"  ⚠️ 找不到 scenes 目录，跳过: {scenes_src_dir}")
                    continue

                # 查找 scenes 目录下的图片
                img_files = []
                patterns = ["*.png", "*.jpg", "*.jpeg", "*.PNG", "*.JPG"]
                for p in patterns:
                    img_files.extend(glob.glob(os.path.join(scenes_src_dir, p)))

                if not img_files:
                    logger.warning(f"  ⚠️ scenes 目录下没有找到任何图片文件，跳过！")
                    continue

                img_files = list(set(img_files)) # 去重
                img_files.sort(key=lambda x: os.path.basename(x))

                # 在 pending_upload 中生成目标文件夹（前缀带上批次号）
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                dest_folder_name = f"{b_id}-{plant_name}_{ts}"
                dest_folder_path = os.path.join(pending_upload_dir, dest_folder_name)
                os.makedirs(dest_folder_path, exist_ok=True)

                # 复制 scenes 下的文件
                logger.info(f"  📂 拷贝 scenes 图片到: pending_upload/{dest_folder_name}")
                for img_path in img_files:
                    shutil.copy2(img_path, os.path.join(dest_folder_path, os.path.basename(img_path)))
                    logger.info(f"    ✓ 拷贝图片: {os.path.basename(img_path)}")

                # 写入 batch_info.json 供后续精准查重使用
                try:
                    batch_info = {
                        "batch_id": b_id,
                        "original_folder": str(f['name']),
                        "plant_name": plant_name,
                        "height": height,
                        "pull_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                    info_path = os.path.join(dest_folder_path, 'batch_info.json')
                    with open(info_path, 'w', encoding='utf-8') as inf:
                        json.dump(batch_info, inf, indent=2, ensure_ascii=False)
                    logger.info("    ✓ 写入批次信息说明文件 batch_info.json")
                except Exception as e:
                    logger.warning(f"  ⚠️ 写入 batch_info.json 失败: {e}")

                # 追加 Excel 记录
                logger.info("  📝 追加记录到 Excel...")
                append_pipeline_excel_record(excel_path, plant_name, height, dest_folder_name)

        if not cancelled_by_user:
            logger.info(f"\n拉取完成！已全部导入到 pending_upload 目录中并生成了 Excel 上架数据记录。")
            input("\n处理完成。按回车键 [Enter] 返回主菜单...")
            return


# ============================================================
# 入口
# ============================================================
if __name__ == '__main__':
    import sys
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    logger.info("=" * 60)
    logger.info("自动上架商品脚本启动")
    logger.info(f"数据库: {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}")
    logger.info("=" * 60)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    pipeline_excel = os.path.join(base_dir, 'pending_upload', '商品上架数据.xlsx')
    template_excel = os.path.join(base_dir, '商品上架模板.xlsx')

    while True:
        # 再次检测，以便用户在中途复制文件后能自动刷新状态
        has_pipeline = os.path.exists(pipeline_excel)
        has_template = os.path.exists(template_excel)

        # 交互式主菜单
        print("\n" + "=" * 60)
        print("  自动商品上架系统主菜单：")
        print("=" * 60)
        print("  [1] 拉取 AI 生图数据 (导入至 pending_upload)")
        print("  [2] 执行森屿家上架 (包含批量上架/单独上架)")
        print("  [q] 退出")
        print("=" * 60)

        choice = input("\n请输入选项 (1/2/q): ").strip().lower()

        if choice == '1':
            pull_ai_generated_data()

        elif choice == '2':
            while True:
                has_pipeline = os.path.exists(pipeline_excel)
                has_template = os.path.exists(template_excel)

                print("\n" + "=" * 60)
                print("  请选择商品上架模式：")
                print("=" * 60)
                if has_pipeline:
                    print("  [1] ai生图批量上架 (数据源: pending_upload/商品上架数据.xlsx)")
                else:
                    print("  [1] ai生图批量上架 (❌ 未检测到 pending_upload/商品上架数据.xlsx)")
                if has_template:
                    print("  [2] 非ai生图上架 (数据源: 商品上架模板.xlsx)")
                else:
                    print("  [2] 非ai生图上架 (❌ 未检测到商品上架模板.xlsx)")
                print("  [b] 返回主菜单")
                print("=" * 60)

                sub_choice = input("\n请输入选项 (1/2/b): ").strip().lower()

                if sub_choice == '1':
                    if not has_pipeline:
                        logger.error("pending_upload/商品上架数据.xlsx 不存在，无法执行批量上架！")
                        input("\n按回车键 [Enter] 返回上级菜单...")
                        continue

                    logger.info(f"已选择：ai生图批量上架模式，数据源: pending_upload/商品上架数据.xlsx")

                    # 直接正式执行
                    logger.info(f"\n{'='*60}")
                    logger.info(">>> 正式执行上架（写入数据库 + 自动归档回写）<<<")
                    logger.info(f"{'='*60}")
                    final_results = process_pipeline_excel(pipeline_excel, dry_run=False)
                    
                    if final_results['total'] == 0:
                        logger.warning("商品上架数据.xlsx 中没有待上架的商品数据，请确认是否填写！")
                    else:
                        logger.info(f"\n最终结果: 成功上架 {final_results['success']} 个，失败 {final_results['failed']} 个")
                    input("\n上架处理完成。按回车键 [Enter] 返回上级菜单...")
                    break

                elif sub_choice == '2':
                    if not has_template:
                        logger.error("商品上架模板.xlsx 不存在，无法执行单独上架！")
                        input("\n按回车键 [Enter] 返回上级菜单...")
                        continue

                    logger.info(f"已选择：非ai生图上架模式，数据源: 商品上架模板.xlsx")
                    logger.info(f"读取 Excel: {template_excel}")
                    products = load_products_from_excel(template_excel)

                    if not products:
                        logger.warning("Excel 中没有商品数据，请先填写模板")
                        input("\n按回车键 [Enter] 返回上级菜单...")
                        continue

                    # 直接正式执行
                    logger.info(f"\n{'='*60}")
                    logger.info(">>> 正式执行上架（写入数据库）<<<")
                    logger.info(f"{'='*60}")
                    results = batch_listing(products, dry_run=False)
                    logger.info(f"\n最终结果: 成功={results['success']}, 失败={results['failed']}")
                    input("\n上架处理完成。按回车键 [Enter] 返回上级菜单...")
                    break

                elif sub_choice == 'b':
                    break
                else:
                    logger.error(f"无效的选项: {sub_choice}，请输入 1、2 或 b")
                    input("\n按回车键 [Enter] 返回...")

        elif choice == 'q':
            logger.info("已退出。")
            sys.exit(0)

        else:
            logger.error(f"无效的选项: {choice}，请输入 1、2 或 q")
            input("\n按回车键 [Enter] 返回主菜单...")

