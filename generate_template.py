# -*- coding: utf-8 -*-
"""生成商品上架 Excel 模板（直观省市多选列）"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def get_all_creators():
    """从数据库中查询所有有效的团长昵称，并加入默认值"""
    creators = []
    try:
        import pymysql
        import sys
        sys.path.append(".")
        from config import DB_CONFIG
        conn = pymysql.connect(**DB_CONFIG)
        cursor = conn.cursor()
        sql = """
            SELECT DISTINCT a.name 
            FROM z100h_wx.wx_user a 
            JOIN guangheshop.litemall_user b ON a.unionid = b.weixin_unionid
            WHERE a.name IS NOT NULL AND a.name != ''
        """
        cursor.execute(sql)
        rows = cursor.fetchall()
        for r in rows:
            if isinstance(r, dict):
                name = r.get('name')
            else:
                name = r[0]
            if name:
                creators.append(name)
        conn.close()
    except Exception as e:
        print(f"[WARNING] 无法连接数据库获取团长列表: {e}")

    creators = list(set(creators))
    for default_name in ["吴鑫霞", "李琳"]:
        if default_name not in creators:
            creators.append(default_name)
    return sorted(creators)


CITY_DATA = {
    "江苏省": [
        ("南京市", 320100), ("无锡市", 320200), ("徐州市", 320300),
        ("常州市", 320400), ("苏州市", 320500), ("南通市", 320600),
        ("连云港市", 320700), ("淮安市", 320800), ("盐城市", 320900),
        ("扬州市", 321000), ("镇江市", 321100), ("泰州市", 321200),
        ("宿迁市", 321300),
    ],
    "浙江省": [
        ("杭州市", 330100), ("宁波市", 330200), ("温州市", 330300),
        ("嘉兴市", 330400), ("湖州市", 330500), ("绍兴市", 330600),
        ("金华市", 330700), ("衢州市", 330800), ("舟山市", 330900),
        ("台州市", 331000), ("丽水市", 331100),
    ],
    "湖北省": [
        ("武汉市", 420100), ("黄石市", 420200), ("十堰市", 420300),
        ("宜昌市", 420500), ("襄阳市", 420600), ("鄂州市", 420700),
        ("荆门市", 420800), ("孝感市", 420900), ("荆州市", 421000),
        ("黄冈市", 421100), ("咸宁市", 421200), ("随州市", 421300),
        ("恩施州", 422800),
    ],
    "广东省": [
        ("广州市", 440100), ("韶关市", 440200), ("深圳市", 440300),
        ("珠海市", 440400), ("汕头市", 440500), ("佛山市", 440600),
        ("江门市", 440700), ("湛江市", 440800), ("茂名市", 440900),
        ("肇庆市", 441200), ("惠州市", 441300), ("梅州市", 441400),
        ("汕尾市", 441500), ("河源市", 441600), ("阳江市", 441700),
        ("清远市", 441800), ("东莞市", 441900), ("中山市", 442000),
        ("潮州市", 445100), ("揭阳市", 445200), ("云浮市", 445300),
    ],
    "上海市": [
        ("上海市", 310100),
    ],
}

# 所有的快捷省份选项
SHORTCUT_OPTIONS = ["江苏省全部", "浙江省全部", "湖北省全部", "广东省全部", "上海市"]

# 所有的城市选项 (去重，保留全名供下拉)
ALL_CITIES = []
for prov, cities in CITY_DATA.items():
    for city_name, _ in cities:
        if city_name not in ALL_CITIES:
            ALL_CITIES.append(city_name)

# 样式常量
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
opt_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
helper_fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type="solid")
data_font = Font(name="微软雅黑", size=10)
desc_font = Font(name="微软雅黑", size=9, color="888888", italic=True)

wb = openpyxl.Workbook()

# ============================
# Sheet 1: 商品上架数据
# ============================
ws = wb.active
ws.title = "商品上架数据"

# 定义列：(列名, 宽度, 是否必填, 说明)
columns = [
    ("商品名称",       20, True,  "如：散尾葵、龟背竹"),
    ("主图",           45, True,  "图片URL 或 本地文件路径"),
    ("轮播图",         55, False, "多张图片用 | 分隔，不填则=主图"),
    ("规格值",         12, True,  "如：45、1盆、大号"),
    ("单位",           8,  False, "默认：盆"),
    ("团长名称",       12, True,  "森屿家用户，如李琳"),
    ("入库数量",       10, True,  "必须 > 0"),
    ("采购单价",       10, True,  "元"),
    ("批发单价",       10, True,  "元"),
    ("省份快捷选择",   15, False, "下拉选择省份（整省全部上架）"),
    ("城市1",          12, False, "下拉选择城市"),
    ("城市2",          12, False, "下拉选择城市"),
    ("城市3",          12, False, "下拉选择城市"),
    ("城市4",          12, False, "下拉选择城市"),
    ("城市5",          12, False, "下拉选择城市"),
    ("手动输入城市",   20, False, "若超5个城市或需手填，用 | 分隔"),
    ("备注",           20, False, "可选"),
]

for col_idx, (name, width, req, desc) in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=f"{name}\n{'(必填)' if req else '(选填)'}")
    cell.font = header_font
    if "城市" in name or "省份" in name:
        cell.fill = helper_fill  # 城市相关列用紫色醒目标记
    else:
        cell.fill = header_fill if req else opt_fill
    cell.alignment = wrap_alignment
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[1].height = 40

for col_idx, (_, _, _, desc) in enumerate(columns, 1):
    cell = ws.cell(row=2, column=col_idx, value=desc)
    cell.font = desc_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 示例数据 (主表直观展示)
sample_data = [
    ["散尾葵", "https://example.com/sanweiqi.jpg", "", "1盆", "盆",
     "李琳", 5, 50.00, 80.00, "", "南京市", "苏州市", "镇江市", "", "", "", "指定三个城市"],
    ["龟背竹", "https://example.com/guibeizhu.jpg", "", "45", "盆",
     "李琳", 3, 30.00, 55.00, "江苏省全部", "", "", "", "", "", "", "整省上架"],
    ["绿萝", "D:\\图片\\lvluo.jpg", "", "小盆", "盆",
     "李琳", 10, 8.00, 15.00, "江苏省全部", "武汉市", "深圳市", "", "", "", "", "混合写法：江苏+武汉+深圳"],
]

for row_idx, row_data in enumerate(sample_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

# ============================
# Sheet 2: 城市与省份下拉源数据 (隐藏列，提供干净环境)
# ============================
ws_list = wb.create_sheet("下拉源数据")

# 写入省份快捷选项在 A 列
ws_list.cell(row=1, column=1, value="省份快捷")
for r_idx, s in enumerate(SHORTCUT_OPTIONS, 2):
    ws_list.cell(row=r_idx, column=1, value=s)

# 写入城市选项在 B 列
ws_list.cell(row=1, column=2, value="所有城市")
for r_idx, c in enumerate(ALL_CITIES, 2):
    ws_list.cell(row=r_idx, column=2, value=c)

# 写入团长选项在 C 列
creators = get_all_creators()
ws_list.cell(row=1, column=3, value="所有团长")
for r_idx, name in enumerate(creators, 2):
    ws_list.cell(row=r_idx, column=3, value=name)

# 隐藏该数据源 sheet，防用户误删
ws_list.views.sheetView[0].showGridLines = True

# ============================
# 设置数据验证 (直接绑定命名范围或单元格范围)
# ============================
# 1. 省份快捷下拉 (J列，第10列)
dv_shortcut = DataValidation(
    type="list",
    formula1=f"=下拉源数据!$A$2:$A${len(SHORTCUT_OPTIONS) + 1}",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="选项无效",
    error="请从下拉列表中选择省份快捷选项",
)
dv_shortcut.promptTitle = "省份全部选择"
dv_shortcut.prompt = "如果整省上架，在此下拉选择省份即可"
dv_shortcut.showInputMessage = True
ws.add_data_validation(dv_shortcut)
dv_shortcut.add("J3:J200")  # 预留到200行

# 2. 城市1-城市5下拉 (K-O列，第11-15列)
dv_cities = DataValidation(
    type="list",
    formula1=f"=下拉源数据!$B$2:$B${len(ALL_CITIES) + 1}",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="选项无效",
    error="请从下拉列表中选择支持的城市名称",
)
dv_cities.promptTitle = "选择可售城市"
dv_cities.prompt = "下拉直接选择对应的可售城市"
dv_cities.showInputMessage = True
ws.add_data_validation(dv_cities)
# 批量绑定到 K-O 列 (11 到 15 列)
for col_idx in range(11, 16):
    col_letter = get_column_letter(col_idx)
    dv_cities.add(f"{col_letter}3:{col_letter}200")

# 3. 团长名称下拉 (F列，第6列)
dv_creator = DataValidation(
    type="list",
    formula1=f"=下拉源数据!$C$2:$C${len(creators) + 1}",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="选项无效",
    error="请选择有效的团长名称",
)
dv_creator.promptTitle = "选择团长"
dv_creator.prompt = "下拉选择商品对应的团长微信昵称"
dv_creator.showInputMessage = True
ws.add_data_validation(dv_creator)
dv_creator.add("F3:F200")

# ============================
# Sheet 3: 填写说明
# ============================
ws2 = wb.create_sheet("填写说明")
instructions = [
    ["字段", "必填", "说明", "示例"],
    ["商品名称", "是", "商品在小程序上显示的名称", "散尾葵"],
    ["主图", "是", "商品主图：URL 或本地路径（自动上传COS）",
     "https://mpfamily-xxx.cos.ap-shanghai.myqcloud.com/activity/xxx.jpg"],
    ["轮播图", "否", "多张用 | 分隔，不填则=主图", "img1.jpg|img2.jpg"],
    ["规格值", "是", "SKU规格描述", "1盆、45、大号"],
    ["单位", "否", '不填默认"盆"', "盆、棵、株"],
    ["团长名称", "是", "森屿家用户，系统自动查询用户ID", "李琳"],
    ["入库数量", "是", "库存数量，必须>0", "5"],
    ["采购单价", "是", "进货成本价（元）", "50.00"],
    ["批发单价", "是", "批发/售卖价（元）", "80.00"],
    ["省份快捷选择", "否", "若整省上架，下拉直接选，如'江苏省全部'", "江苏省全部"],
    ["城市1 ~ 城市5", "否", "若只指定部分城市，可在这5列中直接下拉选择", "南京市、苏州市"],
    ["手动输入城市", "否", "若超过5个城市或需要特殊手填，在此输入，用 | 分隔", "南京|武汉|深圳"],
    ["备注", "否", "批次备注信息", "6月第一批"],
]

for row_idx, row_data in enumerate(instructions, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = data_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 6
ws2.column_dimensions['C'].width = 45
ws2.column_dimensions['D'].width = 35

# 保存
output_path = r"d:\work\自动上架商品\商品上架模板.xlsx"
wb.save(output_path)
print(f"[OK] 模板已生成: {output_path}")
print(f"     设置了「省份快捷选择」列 + 「城市1 ~ 城市5」下拉多列，用户可直接行操作！")
